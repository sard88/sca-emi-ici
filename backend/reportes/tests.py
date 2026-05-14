from datetime import date
from decimal import Decimal
from io import BytesIO
import json
import subprocess
from unittest.mock import patch

from django.contrib import admin
from django.contrib.auth.models import Group
from django.apps import apps
from django.core.exceptions import PermissionDenied
from django.test import RequestFactory, TestCase
from django.utils import timezone

from catalogos.models import Antiguedad, Carrera, GrupoAcademico, Materia, PeriodoEscolar, PlanEstudios, ProgramaAsignatura
from evaluacion.models import (
    Acta,
    CalificacionComponente,
    ComponenteEvaluacion,
    ConformidadDiscente,
    DetalleActa,
    EsquemaEvaluacion,
    ValidacionActa,
)
from relaciones.models import AdscripcionGrupo, AsignacionDocente, Discente, InscripcionMateria, MovimientoAcademico
from trayectoria.models import CatalogoResultadoAcademico, CatalogoSituacionAcademica, EventoSituacionAcademica
from trayectoria.services import obtener_situacion, registrar_baja_temporal, registrar_extraordinario, registrar_reingreso
from usuarios.models import AsignacionCargo, UnidadOrganizacional, Usuario

from .admin import RegistroExportacionAdmin
from .catalogo import CATALOGO_EXPORTACIONES
from .kardex_context import construir_contexto_kardex
from .models import RegistroExportacion
from .services import ServicioExportacion, construir_nombre_archivo
from .utils_calificaciones import calificacion_numerica_con_letra


class ReportesBloque9ATests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.admin_user = Usuario.objects.create_superuser(
            username="admin",
            password="admin",
            nombre_completo="Administrador",
            estado_cuenta=Usuario.ESTADO_ACTIVO,
        )
        self.estadistica = self.crear_usuario_con_grupo("estadistica", "ENCARGADO_ESTADISTICA")
        self.docente = self.crear_usuario_con_grupo("docente", "DOCENTE")
        self.discente = self.crear_usuario_con_grupo("discente", "DISCENTE")
        self.otro_docente = self.crear_usuario_con_grupo("otrodocente", "DOCENTE")

    def crear_usuario_con_grupo(self, username, grupo):
        group, _ = Group.objects.get_or_create(name=grupo)
        user = Usuario.objects.create_user(
            username=username,
            password="admin",
            nombre_completo=username.title(),
            estado_cuenta=Usuario.ESTADO_ACTIVO,
        )
        user.groups.add(group)
        return user

    def crear_registro_directo(self, usuario, tipo=RegistroExportacion.TIPO_ACTA_EVALUACION_PARCIAL):
        return RegistroExportacion.objects.create(
            usuario=usuario,
            tipo_documento=tipo,
            formato=RegistroExportacion.FORMATO_PDF,
            nombre_documento="Documento de prueba",
            nombre_archivo="documento-prueba.pdf",
            estado=RegistroExportacion.ESTADO_SOLICITADA,
        )

    def test_servicio_crea_registro_exportacion_exitoso(self):
        request = self.factory.get("/api/reportes/catalogo/", HTTP_USER_AGENT="TestAgent", REMOTE_ADDR="127.0.0.1")
        request.user = self.admin_user
        registro = ServicioExportacion(self.admin_user, request=request).registrar_solicitud(
            tipo_documento=RegistroExportacion.TIPO_REPORTE_EXPORTACIONES,
            formato=RegistroExportacion.FORMATO_XLSX,
            filtros={"estado": "GENERADA"},
        )

        self.assertEqual(registro.usuario, self.admin_user)
        self.assertEqual(registro.estado, RegistroExportacion.ESTADO_SOLICITADA)
        self.assertEqual(registro.ip_origen, "127.0.0.1")
        self.assertEqual(registro.user_agent, "TestAgent")
        self.assertEqual(registro.filtros_json, {"estado": "GENERADA"})

    def test_servicio_marca_registro_como_generada(self):
        registro = ServicioExportacion(self.admin_user).registrar_solicitud(
            tipo_documento=RegistroExportacion.TIPO_REPORTE_EXPORTACIONES,
            formato=RegistroExportacion.FORMATO_XLSX,
        )
        ServicioExportacion(self.admin_user).marcar_generada(registro, tamano_bytes=128, hash_archivo="abc123")
        registro.refresh_from_db()

        self.assertEqual(registro.estado, RegistroExportacion.ESTADO_GENERADA)
        self.assertEqual(registro.tamano_bytes, 128)
        self.assertEqual(registro.hash_archivo, "abc123")
        self.assertIsNotNone(registro.finalizado_en)

    def test_servicio_marca_registro_como_fallida(self):
        registro = ServicioExportacion(self.admin_user).registrar_solicitud(
            tipo_documento=RegistroExportacion.TIPO_REPORTE_EXPORTACIONES,
            formato=RegistroExportacion.FORMATO_XLSX,
        )
        ServicioExportacion(self.admin_user).marcar_fallida(registro, "Error controlado")
        registro.refresh_from_db()

        self.assertEqual(registro.estado, RegistroExportacion.ESTADO_FALLIDA)
        self.assertIn("Error controlado", registro.mensaje_error)
        self.assertIsNotNone(registro.finalizado_en)

    def test_usuario_autenticado_consulta_solo_sus_exportaciones(self):
        propia = self.crear_registro_directo(self.docente)
        self.crear_registro_directo(self.otro_docente)

        self.client.force_login(self.docente)
        response = self.client.get("/api/exportaciones/")

        self.assertEqual(response.status_code, 200)
        ids = [item["id"] for item in response.json()["items"]]
        self.assertEqual(ids, [propia.id])

    def test_usuario_anonimo_no_consulta_catalogo_ni_exportaciones(self):
        catalogo_response = self.client.get("/api/reportes/catalogo/")
        exportaciones_response = self.client.get("/api/exportaciones/")

        self.assertEqual(catalogo_response.status_code, 401)
        self.assertEqual(exportaciones_response.status_code, 401)

    def test_discente_no_ve_kardex_oficial_como_exportable(self):
        self.client.force_login(self.discente)
        response = self.client.get("/api/reportes/catalogo/")

        self.assertEqual(response.status_code, 200)
        codigos = {item["codigo"] for item in response.json()["items"]}
        self.assertNotIn(RegistroExportacion.TIPO_KARDEX_OFICIAL, codigos)

    def test_docente_no_ve_reportes_globales_como_exportables(self):
        self.client.force_login(self.docente)
        response = self.client.get("/api/reportes/catalogo/")

        self.assertEqual(response.status_code, 200)
        codigos = {item["codigo"] for item in response.json()["items"]}
        self.assertIn(RegistroExportacion.TIPO_ACTA_EVALUACION_PARCIAL, codigos)
        self.assertNotIn(RegistroExportacion.TIPO_REPORTE_DESEMPENO, codigos)
        self.assertNotIn(RegistroExportacion.TIPO_REPORTE_ACTAS_ESTADO, codigos)

    def test_estadistica_ve_catalogo_institucional(self):
        self.client.force_login(self.estadistica)
        response = self.client.get("/api/reportes/catalogo/")

        self.assertEqual(response.status_code, 200)
        codigos = {item["codigo"] for item in response.json()["items"]}
        self.assertIn(RegistroExportacion.TIPO_KARDEX_OFICIAL, codigos)
        self.assertIn(RegistroExportacion.TIPO_REPORTE_DESEMPENO, codigos)
        self.assertIn(RegistroExportacion.TIPO_REPORTE_EXPORTACIONES, codigos)

    def test_admin_ve_catalogo_completo(self):
        self.client.force_login(self.admin_user)
        response = self.client.get("/api/reportes/catalogo/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()["items"]), len(CATALOGO_EXPORTACIONES))

    def test_nombre_archivo_se_normaliza_sin_caracteres_peligrosos(self):
        nombre = construir_nombre_archivo(
            RegistroExportacion.TIPO_ACTA_EVALUACION_PARCIAL,
            RegistroExportacion.FORMATO_PDF,
            objeto_repr="Seguridad Computacional / P1: Niño Águila",
            timestamp="20260513-101500",
        )

        self.assertEqual(nombre, "acta-de-evaluacion-parcial_seguridad-computacional-p1-nino-aguila_20260513-101500.pdf")
        self.assertNotIn("/", nombre)
        self.assertNotIn(" ", nombre)
        self.assertNotIn("ñ", nombre)

    def test_filtros_json_no_guarda_datos_sensibles(self):
        registro = ServicioExportacion(self.admin_user).registrar_solicitud(
            tipo_documento=RegistroExportacion.TIPO_REPORTE_EXPORTACIONES,
            formato=RegistroExportacion.FORMATO_XLSX,
            filtros={
                "periodo": "2026-1",
                "password": "secreto",
                "csrf_token": "abc",
                "nested": {"authorization": "Bearer token", "estado": "ok"},
            },
        )

        self.assertEqual(registro.filtros_json["periodo"], "2026-1")
        self.assertNotIn("password", registro.filtros_json)
        self.assertNotIn("csrf_token", registro.filtros_json)
        self.assertEqual(registro.filtros_json["nested"], {"estado": "ok"})

    def test_admin_registro_exportacion_no_permite_edicion_ordinaria(self):
        model_admin = RegistroExportacionAdmin(RegistroExportacion, admin.site)
        request = self.factory.get("/admin/reportes/registroexportacion/")
        request.user = self.admin_user

        readonly = model_admin.get_readonly_fields(request)
        self.assertIn("filtros_json", readonly)
        self.assertIn("parametros_json", readonly)
        self.assertFalse(model_admin.has_add_permission(request))
        self.assertFalse(model_admin.has_change_permission(request))
        self.assertFalse(model_admin.has_delete_permission(request))
        self.assertNotIn("delete_selected", model_admin.get_actions(request))

    def test_endpoint_prueba_solo_admin_o_estadistica(self):
        self.client.force_login(self.docente)
        denied = self.client.post(
            "/api/exportaciones/registrar-evento-prueba/",
            data={"formato": "PDF"},
            content_type="application/json",
        )
        self.assertEqual(denied.status_code, 403)

        self.client.force_login(self.estadistica)
        allowed = self.client.post(
            "/api/exportaciones/registrar-evento-prueba/",
            data={"formato": "PDF", "filtros": {"token": "no-guardar", "tipo": "prueba"}},
            content_type="application/json",
        )
        self.assertEqual(allowed.status_code, 201)
        registro = RegistroExportacion.objects.get(id=allowed.json()["item"]["id"])
        self.assertEqual(registro.estado, RegistroExportacion.ESTADO_GENERADA)
        self.assertEqual(registro.filtros_json, {"tipo": "prueba"})

    def test_servicio_rechaza_exportacion_no_autorizada_para_docente(self):
        with self.assertRaises(PermissionDenied):
            ServicioExportacion(self.docente).registrar_solicitud(
                tipo_documento=RegistroExportacion.TIPO_REPORTE_DESEMPENO,
                formato=RegistroExportacion.FORMATO_XLSX,
            )


class ReportesBloque9BActasTests(TestCase):
    def setUp(self):
        self.grupo_docente, _ = Group.objects.get_or_create(name="DOCENTE")
        self.grupo_discente, _ = Group.objects.get_or_create(name="DISCENTE")
        self.grupo_estadistica, _ = Group.objects.get_or_create(name="ENCARGADO_ESTADISTICA")
        self.grupo_jefe_carrera, _ = Group.objects.get_or_create(name=AsignacionCargo.ROL_JEFE_SUB_EJEC_CTR)
        self.grupo_jefe_academico, _ = Group.objects.get_or_create(name=AsignacionCargo.ROL_JEFE_ACADEMICO)
        self.admin_user = Usuario.objects.create_superuser(
            username="admin9b",
            password="admin",
            nombre_completo="Administrador 9B",
            estado_cuenta=Usuario.ESTADO_ACTIVO,
        )
        self.estadistica = self.crear_usuario("estadistica9b", "ENCARGADO_ESTADISTICA")
        self.docente = self.crear_usuario(
            "docente9b",
            "DOCENTE",
            nombre="Docente Responsable",
            titulo_profesional="Maestro en ingeniería",
            cedula_profesional="1234567",
        )
        self.otro_docente = self.crear_usuario("otradocente9b", "DOCENTE", nombre="Docente Ajeno")

        self.carrera = Carrera.objects.create(clave="R9BICI", nombre="Ingeniería en Computación e Informática")
        self.crear_jefaturas()
        self.plan = PlanEstudios.objects.create(carrera=self.carrera, clave="R9BPLAN", nombre="Plan 9B")
        self.antiguedad = Antiguedad.objects.create(
            plan_estudios=self.plan,
            clave="R9BANT",
            nombre="Antigüedad 9B",
            anio_inicio=2025,
            anio_fin=2029,
        )
        self.periodo = PeriodoEscolar.objects.create(
            clave="R9BPER",
            anio_escolar="2025-2026",
            periodo_academico=1,
            fecha_inicio=date(2025, 8, 1),
            fecha_fin=date(2026, 1, 31),
        )
        self.grupo = GrupoAcademico.objects.create(
            clave_grupo="R9BG1",
            antiguedad=self.antiguedad,
            periodo=self.periodo,
            semestre_numero=1,
        )
        self.materia = Materia.objects.create(clave="R9BMAT", nombre="Seguridad Computacional", horas_totales=64)
        self.programa = ProgramaAsignatura.objects.create(
            plan_estudios=self.plan,
            materia=self.materia,
            semestre_numero=1,
        )
        self.esquema = EsquemaEvaluacion.objects.create(
            programa_asignatura=self.programa,
            version="v9b",
            num_parciales=EsquemaEvaluacion.PARCIALES_3,
            permite_exencion=True,
        )
        self.componentes = self.crear_componentes()
        self.discentes = [self.crear_discente(idx) for idx in range(1, 4)]
        self.asignacion = AsignacionDocente.objects.create(
            usuario_docente=self.docente,
            grupo_academico=self.grupo,
            programa_asignatura=self.programa,
        )
        self.inscripciones = list(self.asignacion.inscripciones_materia.order_by("discente__matricula"))
        self.acta_p1 = self.crear_acta(
            ComponenteEvaluacion.CORTE_P1,
            [Decimal("5.0"), Decimal("8.0"), Decimal("5.0")],
            estado=Acta.ESTADO_BORRADOR_DOCENTE,
            causas="",
            sugerencias="",
        )
        self.acta_p2 = self.crear_acta(ComponenteEvaluacion.CORTE_P2, [Decimal("7.0"), Decimal("8.0"), Decimal("6.0")])
        self.acta_p3 = self.crear_acta(ComponenteEvaluacion.CORTE_P3, [Decimal("6.0"), Decimal("9.0"), Decimal("6.0")])
        self.acta_final = self.crear_acta(
            ComponenteEvaluacion.CORTE_FINAL,
            [Decimal("8.0"), Decimal("9.0"), Decimal("7.0")],
            estado=Acta.ESTADO_FORMALIZADO_JEFATURA_ACADEMICA,
            formalizada=True,
            causas="N/A",
            sugerencias="N/A",
        )
        for inscripcion, calificacion in zip(self.inscripciones, [Decimal("6.70"), Decimal("8.55"), Decimal("5.95")]):
            inscripcion.calificacion_final = calificacion
            inscripcion.codigo_resultado_oficial = "APROBADO" if calificacion >= Decimal("6.0") else "REPROBADO"
            inscripcion.cerrado_en = timezone.now()
            inscripcion.save(update_fields=["calificacion_final", "codigo_resultado_oficial", "cerrado_en"])

    def crear_usuario(self, username, grupo, nombre=None, **extra_fields):
        user = Usuario.objects.create_user(
            username=username,
            password="admin",
            nombre_completo=nombre or username.title(),
            estado_cuenta=Usuario.ESTADO_ACTIVO,
            **extra_fields,
        )
        user.groups.add(Group.objects.get(name=grupo))
        return user

    def crear_jefaturas(self):
        seccion_academica = UnidadOrganizacional.objects.create(
            clave=UnidadOrganizacional.CLAVE_SECCION_ACADEMICA,
            nombre="Sección Académica",
            tipo_unidad=UnidadOrganizacional.TIPO_SECCION,
        )
        subseccion_carrera = UnidadOrganizacional.objects.create(
            clave="SUB_EJEC_R9B",
            nombre="Subsección de Ejecución y Control ICI",
            tipo_unidad=UnidadOrganizacional.TIPO_SUBSECCION,
            padre=seccion_academica,
            carrera=self.carrera,
        )
        self.jefe_carrera = self.crear_usuario(
            "jefecarrera9b",
            AsignacionCargo.ROL_JEFE_SUB_EJEC_CTR,
            nombre="Jefe Carrera ICI",
        )
        self.jefe_academico = self.crear_usuario(
            "jefeacademico9b",
            AsignacionCargo.ROL_JEFE_ACADEMICO,
            nombre="Jefe Académico",
        )
        AsignacionCargo.objects.create(
            usuario=self.jefe_carrera,
            carrera=self.carrera,
            unidad_organizacional=subseccion_carrera,
            cargo_codigo=AsignacionCargo.CARGO_JEFE_SUB_EJEC_CTR,
            tipo_designacion=AsignacionCargo.DESIGNACION_TITULAR,
        )
        AsignacionCargo.objects.create(
            usuario=self.jefe_academico,
            unidad_organizacional=seccion_academica,
            cargo_codigo=AsignacionCargo.CARGO_JEFE_ACADEMICO,
            tipo_designacion=AsignacionCargo.DESIGNACION_TITULAR,
        )

    def crear_discente(self, idx):
        usuario = self.crear_usuario(f"discente9b{idx}", "DISCENTE", nombre=f"Discente {idx}")
        discente = Discente.objects.create(
            usuario=usuario,
            matricula=f"R9B00{idx}",
            plan_estudios=self.plan,
            antiguedad=self.antiguedad,
        )
        AdscripcionGrupo.objects.create(discente=discente, grupo_academico=self.grupo)
        return discente

    def crear_componentes(self):
        data = [
            (ComponenteEvaluacion.CORTE_P1, "Tareas", Decimal("40.00"), False, 1),
            (ComponenteEvaluacion.CORTE_P1, "Proyecto", Decimal("60.00"), False, 2),
            (ComponenteEvaluacion.CORTE_P2, "Práctica", Decimal("100.00"), False, 1),
            (ComponenteEvaluacion.CORTE_P3, "Participación", Decimal("100.00"), False, 1),
            (ComponenteEvaluacion.CORTE_FINAL, "Trabajo extraclase", Decimal("30.00"), False, 1),
            (ComponenteEvaluacion.CORTE_FINAL, "Examen final", Decimal("70.00"), True, 2),
        ]
        componentes = {}
        for corte, nombre, porcentaje, es_examen, orden in data:
            componentes.setdefault(corte, []).append(
                ComponenteEvaluacion.objects.create(
                    esquema=self.esquema,
                    corte_codigo=corte,
                    nombre=nombre,
                    porcentaje=porcentaje,
                    es_examen=es_examen,
                    orden=orden,
                )
            )
        return componentes

    def crear_acta(self, corte, resultados, *, estado=Acta.ESTADO_FORMALIZADO_JEFATURA_ACADEMICA, formalizada=False, causas="N/A", sugerencias="N/A"):
        acta = Acta.objects.create(
            asignacion_docente=self.asignacion,
            corte_codigo=corte,
            estado_acta=estado,
            esquema=self.esquema,
            esquema_version_snapshot=self.esquema.version,
            peso_parciales_snapshot=self.esquema.peso_parciales,
            peso_final_snapshot=self.esquema.peso_final,
            umbral_exencion_snapshot=self.esquema.umbral_exencion,
            probables_causas_reprobacion=causas,
            sugerencias_academicas=sugerencias,
            creado_por=self.docente,
        )
        if formalizada:
            from django.utils import timezone

            acta.formalizada_en = timezone.now()
            acta.save(update_fields=["formalizada_en"])
        for inscripcion, resultado in zip(self.inscripciones, resultados):
            detalle = DetalleActa.objects.create(
                acta=acta,
                inscripcion_materia=inscripcion,
                resultado_corte=resultado,
                resultado_corte_visible=resultado,
                resultado_final_preliminar=resultado if corte == ComponenteEvaluacion.CORTE_FINAL else None,
                resultado_final_preliminar_visible=resultado if corte == ComponenteEvaluacion.CORTE_FINAL else None,
                resultado_preliminar=DetalleActa.RESULTADO_APROBATORIO if resultado >= Decimal("6.0") else DetalleActa.RESULTADO_REPROBATORIO,
                completo=True,
            )
            self.crear_calificaciones_componentes(detalle, corte, resultado)
        return acta

    def crear_calificaciones_componentes(self, detalle, corte, resultado):
        componentes = self.componentes[corte]
        if corte == ComponenteEvaluacion.CORTE_P1:
            capturados = [resultado, resultado]
        elif corte == ComponenteEvaluacion.CORTE_FINAL:
            capturados = [resultado, resultado]
        else:
            capturados = [resultado]
        for componente, valor_capturado in zip(componentes, capturados):
            CalificacionComponente.objects.create(
                detalle=detalle,
                componente=componente,
                componente_nombre_snapshot=componente.nombre,
                componente_porcentaje_snapshot=componente.porcentaje,
                componente_es_examen_snapshot=componente.es_examen,
                valor_capturado=valor_capturado,
                valor_calculado=valor_capturado * componente.porcentaje / Decimal("100"),
            )

    def abrir_xlsx(self, response):
        from openpyxl import load_workbook

        return load_workbook(BytesIO(response.content), data_only=True)

    def valores_xlsx(self, response):
        wb = self.abrir_xlsx(response)
        ws = wb.active
        return [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]

    def test_exporta_acta_parcial_pdf_con_usuario_autorizado(self):
        self.client.force_login(self.docente)
        with patch("reportes.exportadores.actas_pdf.convertir_xlsx_a_pdf", return_value=b"%PDF-1.4\nmock"):
            response = self.client.get(f"/api/exportaciones/actas/{self.acta_p1.id}/pdf/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))
        self.assertEqual(RegistroExportacion.objects.latest("id").estado, RegistroExportacion.ESTADO_GENERADA)

    def test_exporta_acta_parcial_xlsx_con_usuario_autorizado_y_contenido(self):
        self.client.force_login(self.docente)
        response = self.client.get(f"/api/exportaciones/actas/{self.acta_p1.id}/xlsx/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        valores = self.valores_xlsx(response)
        self.assertIn("DOCUMENTO NO OFICIAL", " ".join(str(valor) for valor in valores))
        self.assertIn("Probables causas de reprobación: N/A", valores)
        self.assertIn("Tareas\n40%", valores)
        self.assertIn(2, valores)
        self.assertIn("Alumnos reprobados:", valores)
        self.assertIn(2, valores)
        self.assertIn("Media:", valores)
        self.assertIn(6, valores)
        self.assertIn("Moda:", valores)
        self.assertIn(5, valores)
        wb = self.abrir_xlsx(response)
        ws = wb.active
        self.assertIn("A1:AX1", {str(rng) for rng in ws.merged_cells.ranges})
        self.assertNotIn("Matrícula", valores)
        self.assertFalse(ws.row_dimensions[18].hidden)
        self.assertTrue(ws.row_dimensions[19].hidden)
        self.assertIn("Maestro en ingeniería", valores)
        self.assertIn("Docente Responsable", " ".join(str(valor) for valor in valores))
        self.assertIn("Cédula profesional: 1234567", valores)
        self.assertIn("Jefe Carrera ICI", " ".join(str(valor) for valor in valores))
        self.assertIn("Jefe Académico", " ".join(str(valor) for valor in valores))

    def test_endpoint_actas_disponibles_docente_lista_solo_propias(self):
        otro_grupo = GrupoAcademico.objects.create(
            clave_grupo="R9BG3",
            antiguedad=self.antiguedad,
            periodo=self.periodo,
            semestre_numero=1,
        )
        otra_asignacion = AsignacionDocente.objects.create(
            usuario_docente=self.otro_docente,
            grupo_academico=otro_grupo,
            programa_asignatura=self.programa,
        )
        acta_ajena = Acta.objects.create(
            asignacion_docente=otra_asignacion,
            corte_codigo=ComponenteEvaluacion.CORTE_P1,
            estado_acta=Acta.ESTADO_BORRADOR_DOCENTE,
            esquema=self.esquema,
            esquema_version_snapshot=self.esquema.version,
            peso_parciales_snapshot=self.esquema.peso_parciales,
            peso_final_snapshot=self.esquema.peso_final,
            umbral_exencion_snapshot=self.esquema.umbral_exencion,
            creado_por=self.otro_docente,
        )

        self.client.force_login(self.docente)
        response = self.client.get("/api/exportaciones/actas-disponibles/")

        self.assertEqual(response.status_code, 200)
        ids = {item["acta_id"] for item in response.json()["items"]}
        self.assertIn(self.acta_p1.id, ids)
        self.assertNotIn(acta_ajena.id, ids)

    def test_endpoint_actas_disponibles_discente_no_lista_actas_completas(self):
        self.client.force_login(self.discentes[0].usuario)
        response = self.client.get("/api/exportaciones/actas-disponibles/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["items"], [])

    def test_endpoint_actas_disponibles_estadistica_lista_actas_y_no_expone_matricula(self):
        self.client.force_login(self.estadistica)
        response = self.client.get("/api/exportaciones/actas-disponibles/")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertGreaterEqual(len(payload["items"]), 4)
        serializado = json.dumps(payload, ensure_ascii=False)
        self.assertIn("url_pdf", payload["items"][0])
        self.assertNotIn("R9B001", serializado)
        self.assertNotIn("matricula", serializado.lower())

    def test_descarga_expone_headers_de_auditoria_para_frontend(self):
        self.client.force_login(self.docente)
        response = self.client.get(
            f"/api/exportaciones/actas/{self.acta_p1.id}/xlsx/",
            HTTP_ORIGIN="http://localhost:3000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["X-Registro-Exportacion-Id"], str(RegistroExportacion.objects.latest("id").id))
        self.assertIn("Content-Disposition", response["Access-Control-Expose-Headers"])
        self.assertIn("X-Registro-Exportacion-Id", response["Access-Control-Expose-Headers"])

    def test_exporta_evaluacion_final_pdf_y_xlsx(self):
        self.client.force_login(self.docente)
        with patch("reportes.exportadores.actas_pdf.convertir_xlsx_a_pdf", return_value=b"%PDF-1.4\nmock"):
            pdf = self.client.get(f"/api/exportaciones/actas/{self.acta_final.id}/pdf/")
        xlsx = self.client.get(f"/api/exportaciones/actas/{self.acta_final.id}/xlsx/")

        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(xlsx.status_code, 200)
        self.assertIn("Evaluación final", self.valores_xlsx(xlsx))
        wb = self.abrir_xlsx(xlsx)
        ws = wb.active
        self.assertIn("A1:H1", {str(rng) for rng in ws.merged_cells.ranges})
        self.assertFalse(ws.row_dimensions[16].hidden)
        self.assertTrue(ws.row_dimensions[17].hidden)

    def test_exporta_calificacion_final_pdf_y_xlsx(self):
        self.client.force_login(self.docente)
        with patch("reportes.exportadores.actas_pdf.convertir_xlsx_a_pdf", return_value=b"%PDF-1.4\nmock"):
            pdf = self.client.get(f"/api/exportaciones/asignaciones/{self.asignacion.id}/calificacion-final/pdf/")
        xlsx = self.client.get(f"/api/exportaciones/asignaciones/{self.asignacion.id}/calificacion-final/xlsx/")

        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(xlsx.status_code, 200)
        valores = self.valores_xlsx(xlsx)
        self.assertIn("Parcial\n1", valores)
        self.assertIn("Parcial \n2 ", valores)
        self.assertIn("Parcial \n3 ", valores)
        self.assertIn("Evaluación Final", valores)
        self.assertIn("Calificación final", valores)
        wb = self.abrir_xlsx(xlsx)
        ws = wb.active
        self.assertIn("A1:I1", {str(rng) for rng in ws.merged_cells.ranges})
        self.assertFalse(ws.row_dimensions[16].hidden)
        self.assertTrue(ws.row_dimensions[17].hidden)

    def test_docente_no_puede_exportar_acta_ajena(self):
        otro_grupo = GrupoAcademico.objects.create(
            clave_grupo="R9BG2",
            antiguedad=self.antiguedad,
            periodo=self.periodo,
            semestre_numero=1,
        )
        otra_asignacion = AsignacionDocente.objects.create(
            usuario_docente=self.otro_docente,
            grupo_academico=otro_grupo,
            programa_asignatura=self.programa,
        )
        acta_ajena = Acta.objects.create(
            asignacion_docente=otra_asignacion,
            corte_codigo=ComponenteEvaluacion.CORTE_P1,
            estado_acta=Acta.ESTADO_BORRADOR_DOCENTE,
            esquema=self.esquema,
            esquema_version_snapshot=self.esquema.version,
            peso_parciales_snapshot=self.esquema.peso_parciales,
            peso_final_snapshot=self.esquema.peso_final,
            umbral_exencion_snapshot=self.esquema.umbral_exencion,
            creado_por=self.otro_docente,
        )

        self.client.force_login(self.docente)
        response = self.client.get(f"/api/exportaciones/actas/{acta_ajena.id}/pdf/")

        self.assertEqual(response.status_code, 403)

    def test_discente_no_puede_exportar_acta_completa(self):
        self.client.force_login(self.discentes[0].usuario)
        response = self.client.get(f"/api/exportaciones/actas/{self.acta_p1.id}/pdf/")

        self.assertEqual(response.status_code, 403)

    def test_estadistica_puede_exportar_acta_formalizada(self):
        self.client.force_login(self.estadistica)
        with patch("reportes.exportadores.actas_pdf.convertir_xlsx_a_pdf", return_value=b"%PDF-1.4\nmock"):
            response = self.client.get(f"/api/exportaciones/actas/{self.acta_final.id}/pdf/")

        self.assertEqual(response.status_code, 200)

    def test_exportacion_fallida_crea_registro_fallida(self):
        self.client.force_login(self.docente)
        with patch("reportes.exportadores.actas_pdf.convertir_xlsx_a_pdf", side_effect=RuntimeError("fallo controlado")):
            response = self.client.get(f"/api/exportaciones/actas/{self.acta_p1.id}/pdf/")

        self.assertEqual(response.status_code, 500)
        registro = RegistroExportacion.objects.latest("id")
        self.assertEqual(registro.estado, RegistroExportacion.ESTADO_FALLIDA)
        self.assertIn("fallo controlado", registro.mensaje_error)

    def test_nombre_archivo_no_contiene_datos_sensibles(self):
        self.client.force_login(self.docente)
        with patch("reportes.exportadores.actas_pdf.convertir_xlsx_a_pdf", return_value=b"%PDF-1.4\nmock"):
            response = self.client.get(f"/api/exportaciones/actas/{self.acta_p1.id}/pdf/")

        disposition = response["Content-Disposition"]
        filename = disposition.split('filename="', 1)[1].rstrip('"')
        self.assertNotIn("Discente", disposition)
        self.assertNotIn("R9B001", disposition)
        self.assertNotIn(" ", filename)

    def test_acta_formalizada_no_muestra_marca_borrador(self):
        self.client.force_login(self.docente)
        response = self.client.get(f"/api/exportaciones/actas/{self.acta_final.id}/xlsx/")

        valores = " ".join(str(valor) for valor in self.valores_xlsx(response))
        self.assertNotIn("DOCUMENTO NO OFICIAL", valores)
        self.assertNotIn("BORRADOR", valores)

    def test_calificacion_final_sin_acta_final_sale_no_oficial(self):
        self.acta_final.estado_acta = Acta.ESTADO_ARCHIVADO
        self.acta_final.save(update_fields=["estado_acta"])

        self.client.force_login(self.docente)
        response = self.client.get(f"/api/exportaciones/asignaciones/{self.asignacion.id}/calificacion-final/xlsx/")

        self.assertEqual(response.status_code, 200)
        valores = " ".join(str(valor) for valor in self.valores_xlsx(response))
        self.assertIn("DOCUMENTO NO OFICIAL", valores)

    def test_exportar_no_modifica_acta_ni_inscripcion(self):
        acta_estado = self.acta_p1.estado_acta
        inscripcion_valor = self.inscripciones[0].calificacion_final

        self.client.force_login(self.docente)
        response = self.client.get(f"/api/exportaciones/actas/{self.acta_p1.id}/xlsx/")

        self.assertEqual(response.status_code, 200)
        self.acta_p1.refresh_from_db()
        self.inscripciones[0].refresh_from_db()
        self.assertEqual(self.acta_p1.estado_acta, acta_estado)
        self.assertEqual(self.inscripciones[0].calificacion_final, inscripcion_valor)

    def test_conversion_pdf_usa_libreoffice_headless(self):
        from reportes.exportadores.actas_pdf import convertir_xlsx_a_pdf

        def fake_run(cmd, check, stdout, stderr, timeout, env):
            outdir = cmd[cmd.index("--outdir") + 1]
            pdf_path = f"{outdir}/acta.pdf"
            with open(pdf_path, "wb") as file:
                file.write(b"%PDF-1.4\nmock")
            return subprocess.CompletedProcess(cmd, 0, b"", b"")

        with patch("reportes.exportadores.libreoffice_utils.libreoffice_binary", return_value="soffice"), patch(
            "reportes.exportadores.libreoffice_utils.subprocess.run",
            side_effect=fake_run,
        ) as run_mock:
            contenido = convertir_xlsx_a_pdf(b"contenido xlsx")

        self.assertTrue(contenido.startswith(b"%PDF"))
        comando = run_mock.call_args.args[0]
        self.assertIn("--headless", comando)
        self.assertIn("--convert-to", comando)
        self.assertIn("pdf", comando)

    def test_exporta_kardex_pdf_como_estadistica(self):
        self.client.force_login(self.estadistica)
        with patch("reportes.exportadores.kardex_pdf.convertir_xlsx_a_pdf", return_value=b"%PDF-1.4\nkardex"):
            response = self.client.get(f"/api/exportaciones/kardex/{self.discentes[0].id}/pdf/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertEqual(response["X-Registro-Exportacion-Id"], str(RegistroExportacion.objects.latest("id").id))
        registro = RegistroExportacion.objects.latest("id")
        self.assertEqual(registro.tipo_documento, RegistroExportacion.TIPO_KARDEX_OFICIAL)
        self.assertEqual(registro.formato, RegistroExportacion.FORMATO_PDF)
        self.assertEqual(registro.estado, RegistroExportacion.ESTADO_GENERADA)

    def test_exporta_kardex_pdf_como_admin(self):
        self.client.force_login(self.admin_user)
        with patch("reportes.exportadores.kardex_pdf.convertir_xlsx_a_pdf", return_value=b"%PDF-1.4\nkardex"):
            response = self.client.get(f"/api/exportaciones/kardex/{self.discentes[0].id}/pdf/")

        self.assertEqual(response.status_code, 200)

    def test_jefatura_carrera_exporta_kardex_de_su_ambito(self):
        self.client.force_login(self.jefe_carrera)
        with patch("reportes.exportadores.kardex_pdf.convertir_xlsx_a_pdf", return_value=b"%PDF-1.4\nkardex"):
            response = self.client.get(f"/api/exportaciones/kardex/{self.discentes[0].id}/pdf/")

        self.assertEqual(response.status_code, 200)

    def test_jefatura_carrera_no_exporta_kardex_fuera_de_ambito(self):
        otra_carrera = Carrera.objects.create(clave="R9BOTRA", nombre="Otra carrera")
        otro_plan = PlanEstudios.objects.create(carrera=otra_carrera, clave="R9BOPLAN", nombre="Plan externo")
        otra_antiguedad = Antiguedad.objects.create(
            plan_estudios=otro_plan,
            clave="R9BOANT",
            nombre="Antigüedad externa",
            anio_inicio=2025,
            anio_fin=2029,
        )
        usuario_discente = self.crear_usuario("discenteexterno9c", "DISCENTE", nombre="Discente Externo")
        discente_externo = Discente.objects.create(
            usuario=usuario_discente,
            matricula="R9BEXT",
            plan_estudios=otro_plan,
            antiguedad=otra_antiguedad,
        )

        self.client.force_login(self.jefe_carrera)
        response = self.client.get(f"/api/exportaciones/kardex/{discente_externo.id}/pdf/")

        self.assertEqual(response.status_code, 403)

    def test_discente_y_docente_no_exportan_kardex(self):
        self.client.force_login(self.discentes[0].usuario)
        response_discente = self.client.get(f"/api/exportaciones/kardex/{self.discentes[0].id}/pdf/")

        self.client.force_login(self.docente)
        response_docente = self.client.get(f"/api/exportaciones/kardex/{self.discentes[0].id}/pdf/")

        self.assertEqual(response_discente.status_code, 403)
        self.assertEqual(response_docente.status_code, 403)

    def test_exportacion_kardex_fallida_registra_fallida(self):
        self.client.force_login(self.estadistica)
        with patch("reportes.exportadores.kardex_pdf.convertir_xlsx_a_pdf", side_effect=RuntimeError("fallo kardex")):
            response = self.client.get(f"/api/exportaciones/kardex/{self.discentes[0].id}/pdf/")

        self.assertEqual(response.status_code, 500)
        registro = RegistroExportacion.objects.latest("id")
        self.assertEqual(registro.tipo_documento, RegistroExportacion.TIPO_KARDEX_OFICIAL)
        self.assertEqual(registro.estado, RegistroExportacion.ESTADO_FALLIDA)
        self.assertIn("fallo kardex", registro.mensaje_error)

    def test_nombre_archivo_kardex_no_contiene_datos_sensibles(self):
        self.client.force_login(self.estadistica)
        with patch("reportes.exportadores.kardex_pdf.convertir_xlsx_a_pdf", return_value=b"%PDF-1.4\nkardex"):
            response = self.client.get(f"/api/exportaciones/kardex/{self.discentes[0].id}/pdf/")

        disposition = response["Content-Disposition"]
        self.assertNotIn(self.discentes[0].usuario.nombre_visible, disposition)
        self.assertNotIn(self.discentes[0].matricula, disposition)
        self.assertIn(f"discente-{self.discentes[0].id}", disposition)

    def test_catalogo_marca_kardex_pdf_implementado_para_estadistica_y_no_discente(self):
        self.client.force_login(self.estadistica)
        response_estadistica = self.client.get("/api/reportes/catalogo/")
        item = next(item for item in response_estadistica.json()["items"] if item["codigo"] == RegistroExportacion.TIPO_KARDEX_OFICIAL)

        self.assertTrue(item["implementado"])
        self.assertTrue(item["disponible"])
        self.assertEqual(item["formatos_soportados"], [RegistroExportacion.FORMATO_PDF])

        self.client.force_login(self.discentes[0].usuario)
        response_discente = self.client.get("/api/reportes/catalogo/")
        codigos = {item["codigo"] for item in response_discente.json()["items"]}
        self.assertNotIn(RegistroExportacion.TIPO_KARDEX_OFICIAL, codigos)

    def test_kardex_disponibles_requiere_autenticacion(self):
        response = self.client.get("/api/exportaciones/kardex-disponibles/")

        self.assertEqual(response.status_code, 401)

    def test_kardex_disponibles_admin_y_estadistica_ven_resultados_sin_matricula(self):
        for usuario in [self.admin_user, self.estadistica]:
            self.client.force_login(usuario)
            response = self.client.get("/api/exportaciones/kardex-disponibles/")

            self.assertEqual(response.status_code, 200)
            data = response.json()["items"]
            self.assertGreaterEqual(len(data), len(self.discentes))
            primer_item = data[0]
            self.assertIn("url_kardex_pdf", primer_item)
            self.assertIn("discente_id", primer_item)
            self.assertNotIn("matricula", primer_item)
            self.assertNotIn(self.discentes[0].matricula, json.dumps(data))

    def test_kardex_disponibles_jefatura_carrera_solo_ambito(self):
        otra_carrera = Carrera.objects.create(clave="R9B10C2", nombre="Otra carrera 10C2")
        otro_plan = PlanEstudios.objects.create(carrera=otra_carrera, clave="R9B10C2P", nombre="Plan externo 10C2")
        otra_antiguedad = Antiguedad.objects.create(
            plan_estudios=otro_plan,
            clave="R9B10C2A",
            nombre="Antigüedad externa 10C2",
            anio_inicio=2025,
            anio_fin=2029,
        )
        usuario_discente = self.crear_usuario("discenteexterno10c2", "DISCENTE", nombre="Discente Externo 10C2")
        discente_externo = Discente.objects.create(
            usuario=usuario_discente,
            matricula="R9B10C2EXT",
            plan_estudios=otro_plan,
            antiguedad=otra_antiguedad,
        )

        self.client.force_login(self.jefe_carrera)
        response = self.client.get("/api/exportaciones/kardex-disponibles/")

        self.assertEqual(response.status_code, 200)
        ids = {item["discente_id"] for item in response.json()["items"]}
        self.assertIn(self.discentes[0].id, ids)
        self.assertNotIn(discente_externo.id, ids)

    def test_kardex_disponibles_discente_y_docente_no_ven_resultados(self):
        self.client.force_login(self.discentes[0].usuario)
        response_discente = self.client.get("/api/exportaciones/kardex-disponibles/")

        self.client.force_login(self.docente)
        response_docente = self.client.get("/api/exportaciones/kardex-disponibles/")

        self.assertEqual(response_discente.status_code, 200)
        self.assertEqual(response_discente.json()["items"], [])
        self.assertEqual(response_docente.status_code, 200)
        self.assertEqual(response_docente.json()["items"], [])

    def test_kardex_disponibles_respeta_filtros(self):
        self.client.force_login(self.estadistica)
        response = self.client.get(
            "/api/exportaciones/kardex-disponibles/",
            {"q": self.discentes[0].usuario.nombre_visible.split()[0], "carrera": self.carrera.clave, "situacion": Discente.SITUACION_REGULAR},
        )

        self.assertEqual(response.status_code, 200)
        ids = {item["discente_id"] for item in response.json()["items"]}
        self.assertIn(self.discentes[0].id, ids)

    def test_kardex_no_crea_modelo_transaccional(self):
        nombres_modelos = {model.__name__ for model in apps.get_models()}

        self.assertNotIn("KardexOficial", nombres_modelos)

    def test_exportar_kardex_no_modifica_acta_ni_inscripcion(self):
        estado_acta = self.acta_final.estado_acta
        calificacion = self.inscripciones[0].calificacion_final

        self.client.force_login(self.estadistica)
        with patch("reportes.exportadores.kardex_pdf.convertir_xlsx_a_pdf", return_value=b"%PDF-1.4\nkardex"):
            response = self.client.get(f"/api/exportaciones/kardex/{self.discentes[0].id}/pdf/")

        self.assertEqual(response.status_code, 200)
        self.acta_final.refresh_from_db()
        self.inscripciones[0].refresh_from_db()
        self.assertEqual(self.acta_final.estado_acta, estado_acta)
        self.assertEqual(self.inscripciones[0].calificacion_final, calificacion)

    def test_contexto_kardex_muestra_ee_y_excluye_no_numericas_del_promedio(self):
        self.inscripciones[0].calificacion_final = Decimal("5.0")
        self.inscripciones[0].codigo_resultado_oficial = CatalogoResultadoAcademico.CLAVE_REPROBADO
        self.inscripciones[0].cerrado_en = timezone.now()
        self.inscripciones[0].save(update_fields=["calificacion_final", "codigo_resultado_oficial", "cerrado_en"])
        registrar_extraordinario(
            self.inscripciones[0],
            Decimal("8.0"),
            timezone.localdate(),
            self.estadistica,
        )

        materia_no_numerica = Materia.objects.create(
            clave="R9BACR",
            nombre="Formación institucional acreditable",
            horas_totales=32,
        )
        programa_no_numerico = ProgramaAsignatura.objects.create(
            plan_estudios=self.plan,
            materia=materia_no_numerica,
            semestre_numero=1,
        )
        esquema_no_numerico = EsquemaEvaluacion.objects.create(
            programa_asignatura=programa_no_numerico,
            version="v9b-acre",
            num_parciales=EsquemaEvaluacion.PARCIALES_1,
            permite_exencion=False,
        )
        asignacion_no_numerica = AsignacionDocente.objects.create(
            usuario_docente=self.docente,
            grupo_academico=self.grupo,
            programa_asignatura=programa_no_numerico,
        )
        inscripcion_no_numerica = asignacion_no_numerica.inscripciones_materia.get(discente=self.discentes[0])
        acta_no_numerica = Acta.objects.create(
            asignacion_docente=asignacion_no_numerica,
            corte_codigo=ComponenteEvaluacion.CORTE_FINAL,
            estado_acta=Acta.ESTADO_FORMALIZADO_JEFATURA_ACADEMICA,
            esquema=esquema_no_numerico,
            esquema_version_snapshot=esquema_no_numerico.version,
            peso_parciales_snapshot=esquema_no_numerico.peso_parciales,
            peso_final_snapshot=esquema_no_numerico.peso_final,
            umbral_exencion_snapshot=esquema_no_numerico.umbral_exencion,
            creado_por=self.docente,
        )
        DetalleActa.objects.create(
            acta=acta_no_numerica,
            inscripcion_materia=inscripcion_no_numerica,
            resultado_preliminar=DetalleActa.RESULTADO_APROBATORIO,
            completo=True,
        )
        inscripcion_no_numerica.calificacion_final = None
        inscripcion_no_numerica.codigo_resultado_oficial = CatalogoResultadoAcademico.CLAVE_ACREDITADA
        inscripcion_no_numerica.cerrado_en = timezone.now()
        inscripcion_no_numerica.save(
            update_fields=["calificacion_final", "codigo_resultado_oficial", "cerrado_en"]
        )

        contexto = construir_contexto_kardex(self.discentes[0])
        materias = [materia for anio in contexto.anios for materia in anio.materias]
        materia_no_numerica_contexto = next(materia for materia in materias if materia.resultado_display == "ACREDITADA")
        materia_ee = next(materia for materia in materias if materia.marca == "EE")

        self.assertFalse(materia_no_numerica_contexto.es_numerica)
        self.assertEqual(materia_no_numerica_contexto.calificacion_display, "ACREDITADA")
        self.assertEqual(materia_ee.calificacion_display, "8.0")
        self.assertEqual(contexto.anios[0].promedio_anual_display, "8.0")

    def test_calificacion_con_letra_documental(self):
        self.assertEqual(calificacion_numerica_con_letra(Decimal("10.0")), "10.0 (DIEZ PUNTO CERO)")
        self.assertEqual(calificacion_numerica_con_letra(Decimal("9.5")), "9.5 (NUEVE PUNTO CINCO)")
        self.assertEqual(calificacion_numerica_con_letra(Decimal("8.0")), "8.0 (OCHO PUNTO CERO)")

    def crear_validaciones_operativas(self):
        cargo_carrera = AsignacionCargo.objects.get(usuario=self.jefe_carrera)
        cargo_academico = AsignacionCargo.objects.get(usuario=self.jefe_academico)
        return (
            ValidacionActa.objects.create(
                acta=self.acta_p2,
                asignacion_cargo=cargo_carrera,
                usuario=self.jefe_carrera,
                etapa_validacion=ValidacionActa.ETAPA_JEFATURA_CARRERA,
                accion=ValidacionActa.ACCION_VALIDA,
                ip_origen="127.0.0.10",
                comentario="Validación de carrera",
            ),
            ValidacionActa.objects.create(
                acta=self.acta_p2,
                asignacion_cargo=cargo_academico,
                usuario=self.jefe_academico,
                etapa_validacion=ValidacionActa.ETAPA_JEFATURA_ACADEMICA,
                accion=ValidacionActa.ACCION_FORMALIZA,
                ip_origen="127.0.0.11",
                comentario="Formalización académica",
            ),
        )

    def preparar_actas_pendientes(self):
        ahora = timezone.now()
        self.acta_p1.estado_acta = Acta.ESTADO_REMITIDO_JEFATURA_CARRERA
        self.acta_p1.remitida_en = ahora
        self.acta_p1.publicada_en = ahora
        self.acta_p1.save(update_fields=["estado_acta", "remitida_en", "publicada_en"])
        self.acta_p3.estado_acta = Acta.ESTADO_VALIDADO_JEFATURA_CARRERA
        self.acta_p3.remitida_en = ahora
        self.acta_p3.save(update_fields=["estado_acta", "remitida_en"])

    def crear_inconformidad_operativa(self):
        detalle = self.acta_p2.detalles.select_related("inscripcion_materia__discente__usuario").first()
        return ConformidadDiscente.objects.create(
            detalle=detalle,
            discente=detalle.inscripcion_materia.discente,
            usuario=detalle.inscripcion_materia.discente.usuario,
            estado_conformidad=ConformidadDiscente.ESTADO_INCONFORME,
            comentario="Solicito revisión del corte.",
        )

    def crear_acta_externa_operativa(self):
        otra_carrera = Carrera.objects.create(clave="R9FEXT", nombre="Carrera externa 9F")
        otro_plan = PlanEstudios.objects.create(carrera=otra_carrera, clave="R9FPLAN", nombre="Plan externo 9F")
        otra_antiguedad = Antiguedad.objects.create(
            plan_estudios=otro_plan,
            clave="R9FANT",
            nombre="Antigüedad externa 9F",
            anio_inicio=2025,
            anio_fin=2029,
        )
        otro_grupo = GrupoAcademico.objects.create(
            clave_grupo="R9FEXTG",
            antiguedad=otra_antiguedad,
            periodo=self.periodo,
            semestre_numero=1,
        )
        otra_materia = Materia.objects.create(clave="R9FEXTM", nombre="Materia externa 9F", horas_totales=40)
        otro_programa = ProgramaAsignatura.objects.create(
            plan_estudios=otro_plan,
            materia=otra_materia,
            semestre_numero=1,
        )
        otro_esquema = EsquemaEvaluacion.objects.create(
            programa_asignatura=otro_programa,
            version="v9f-ext",
            num_parciales=EsquemaEvaluacion.PARCIALES_1,
        )
        ComponenteEvaluacion.objects.create(
            esquema=otro_esquema,
            corte_codigo=ComponenteEvaluacion.CORTE_P1,
            nombre="Componente externo",
            porcentaje=Decimal("100.00"),
            orden=1,
        )
        ComponenteEvaluacion.objects.create(
            esquema=otro_esquema,
            corte_codigo=ComponenteEvaluacion.CORTE_FINAL,
            nombre="Examen externo",
            porcentaje=Decimal("100.00"),
            es_examen=True,
            orden=1,
        )
        otra_asignacion = AsignacionDocente.objects.create(
            usuario_docente=self.otro_docente,
            grupo_academico=otro_grupo,
            programa_asignatura=otro_programa,
        )
        return Acta.objects.create(
            asignacion_docente=otra_asignacion,
            corte_codigo=ComponenteEvaluacion.CORTE_P1,
            estado_acta=Acta.ESTADO_BORRADOR_DOCENTE,
            esquema=otro_esquema,
            esquema_version_snapshot=otro_esquema.version,
            peso_parciales_snapshot=otro_esquema.peso_parciales,
            peso_final_snapshot=otro_esquema.peso_final,
            umbral_exencion_snapshot=otro_esquema.umbral_exencion,
            creado_por=self.otro_docente,
        )

    def assert_xlsx_response(self, response):
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(response.content.startswith(b"PK"))
        self.assertIn("X-Registro-Exportacion-Id", response)

    def test_reportes_operativos_json_y_xlsx_como_estadistica(self):
        self.preparar_actas_pendientes()
        self.crear_inconformidad_operativa()
        self.crear_validaciones_operativas()
        RegistroExportacion.objects.create(
            usuario=self.docente,
            tipo_documento=RegistroExportacion.TIPO_ACTA_EVALUACION_PARCIAL,
            formato=RegistroExportacion.FORMATO_XLSX,
            nombre_documento="Acta previa",
            nombre_archivo="acta-previa.xlsx",
            estado=RegistroExportacion.ESTADO_GENERADA,
        )

        endpoints = [
            "actas-estado",
            "actas-pendientes",
            "inconformidades",
            "sin-conformidad",
            "actas-formalizadas",
            "validaciones-acta",
            "exportaciones-realizadas",
        ]
        self.client.force_login(self.estadistica)
        for slug in endpoints:
            json_response = self.client.get(f"/api/reportes/operativos/{slug}/")
            self.assertEqual(json_response.status_code, 200, slug)
            self.assertIn("total", json_response.json(), slug)
            self.assertIn("columnas", json_response.json(), slug)

            xlsx_response = self.client.get(f"/api/exportaciones/reportes/{slug}/xlsx/")
            self.assert_xlsx_response(xlsx_response)
            registro = RegistroExportacion.objects.get(id=xlsx_response["X-Registro-Exportacion-Id"])
            self.assertEqual(registro.estado, RegistroExportacion.ESTADO_GENERADA)
            self.assertEqual(registro.formato, RegistroExportacion.FORMATO_XLSX)

    def test_reportes_operativos_bloquean_anonimo_discente_y_docente_global(self):
        anonimo = self.client.get("/api/reportes/operativos/actas-estado/")
        self.assertEqual(anonimo.status_code, 401)

        self.client.force_login(self.discentes[0].usuario)
        discente_json = self.client.get("/api/reportes/operativos/actas-estado/")
        discente_xlsx = self.client.get("/api/exportaciones/reportes/actas-estado/xlsx/")
        self.assertEqual(discente_json.status_code, 403)
        self.assertEqual(discente_xlsx.status_code, 403)

        self.client.force_login(self.docente)
        docente_json = self.client.get("/api/reportes/operativos/actas-estado/")
        docente_xlsx = self.client.get("/api/exportaciones/reportes/actas-estado/xlsx/")
        self.assertEqual(docente_json.status_code, 403)
        self.assertEqual(docente_xlsx.status_code, 403)

    def test_reportes_operativos_jefatura_carrera_filtra_por_ambito(self):
        acta_externa = self.crear_acta_externa_operativa()

        self.client.force_login(self.jefe_carrera)
        response = self.client.get("/api/reportes/operativos/actas-estado/")

        self.assertEqual(response.status_code, 200)
        ids = {item["acta_id"] for item in response.json()["items"]}
        self.assertIn(self.acta_p1.id, ids)
        self.assertNotIn(acta_externa.id, ids)

    def test_reporte_inconformidades_incluye_comentario_y_no_matricula(self):
        self.crear_inconformidad_operativa()

        self.client.force_login(self.estadistica)
        response = self.client.get("/api/reportes/operativos/inconformidades/")

        self.assertEqual(response.status_code, 200)
        serializado = json.dumps(response.json(), ensure_ascii=False)
        self.assertIn("Solicito revisión del corte.", serializado)
        self.assertNotIn(self.discentes[0].matricula, serializado)
        self.assertNotIn("matricula", serializado.lower())

    def test_reporte_actas_sin_conformidad_lista_excepciones_no_bloqueantes(self):
        self.acta_p1.estado_acta = Acta.ESTADO_PUBLICADO_DISCENTE
        self.acta_p1.publicada_en = timezone.now()
        self.acta_p1.save(update_fields=["estado_acta", "publicada_en"])

        self.client.force_login(self.estadistica)
        response = self.client.get("/api/reportes/operativos/sin-conformidad/")

        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(response.json()["total"], len(self.discentes))
        self.assertIn("Sin conformidad vigente registrada", json.dumps(response.json(), ensure_ascii=False))

    def test_reporte_validaciones_acta_muestra_trazabilidad(self):
        self.crear_validaciones_operativas()

        self.client.force_login(self.estadistica)
        response = self.client.get("/api/reportes/operativos/validaciones-acta/")

        self.assertEqual(response.status_code, 200)
        payload = json.dumps(response.json(), ensure_ascii=False)
        self.assertIn("Validación de carrera", payload)
        self.assertIn("Formalización académica", payload)
        self.assertIn("Jefe Carrera ICI", payload)

    def test_reporte_exportaciones_realizadas_admin_ve_registros(self):
        registro = RegistroExportacion.objects.create(
            usuario=self.docente,
            tipo_documento=RegistroExportacion.TIPO_ACTA_EVALUACION_PARCIAL,
            formato=RegistroExportacion.FORMATO_PDF,
            nombre_documento="Acta exportada",
            nombre_archivo="acta-exportada.pdf",
            estado=RegistroExportacion.ESTADO_GENERADA,
            hash_archivo="abc123",
        )

        self.client.force_login(self.admin_user)
        response = self.client.get("/api/reportes/operativos/exportaciones-realizadas/")

        self.assertEqual(response.status_code, 200)
        ids = {item["registro_id"] for item in response.json()["items"]}
        self.assertIn(registro.id, ids)

    def test_exportacion_reporte_operativo_fallida_registra_fallida(self):
        self.client.force_login(self.estadistica)
        with patch("reportes.reportes_operativos.generar_reporte_xlsx", side_effect=RuntimeError("fallo reporte")):
            response = self.client.get("/api/exportaciones/reportes/actas-estado/xlsx/")

        self.assertEqual(response.status_code, 500)
        registro = RegistroExportacion.objects.latest("id")
        self.assertEqual(registro.tipo_documento, RegistroExportacion.TIPO_REPORTE_ACTAS_ESTADO)
        self.assertEqual(registro.estado, RegistroExportacion.ESTADO_FALLIDA)
        self.assertIn("fallo reporte", registro.mensaje_error)

    def test_exportacion_reporte_operativo_nombre_seguro_y_filtros_sanitizados(self):
        self.client.force_login(self.estadistica)
        response = self.client.get(
            "/api/exportaciones/reportes/actas-estado/xlsx/",
            {"periodo": self.periodo.clave, "carrera": self.carrera.clave, "password": "secreto"},
        )

        self.assert_xlsx_response(response)
        disposition = response["Content-Disposition"]
        filename = disposition.split('filename="', 1)[1].rstrip('"')
        registro = RegistroExportacion.objects.get(id=response["X-Registro-Exportacion-Id"])
        self.assertNotIn(" ", filename)
        self.assertNotIn(self.discentes[0].matricula, filename)
        self.assertEqual(registro.filtros_json, {"periodo": self.periodo.clave, "carrera": self.carrera.clave})

    def marcar_exento_final(self):
        detalle = self.acta_final.detalles.order_by("id").first()
        detalle.exencion_aplica = True
        detalle.promedio_parciales = Decimal("9.4")
        detalle.promedio_parciales_visible = Decimal("9.4")
        detalle.save(update_fields=["exencion_aplica", "promedio_parciales", "promedio_parciales_visible"])
        componente = detalle.calificaciones_componentes.filter(componente_es_examen_snapshot=True).first()
        componente.sustituido_por_exencion = True
        componente.valor_calculado = Decimal("6.58")
        componente.save(update_fields=["sustituido_por_exencion", "valor_calculado"])

    def assert_xlsx_response_desempeno(self, response):
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(response.content.startswith(b"PK"))
        self.assertIn("X-Registro-Exportacion-Id", response)

    def test_reportes_desempeno_json_y_xlsx_como_estadistica(self):
        self.marcar_exento_final()
        endpoints = [
            ("aprobados-reprobados", "/api/reportes/desempeno/aprobados-reprobados/", "/api/exportaciones/reportes/aprobados-reprobados/xlsx/"),
            ("promedios", "/api/reportes/desempeno/promedios/", "/api/exportaciones/reportes/promedios/xlsx/"),
            ("distribucion", "/api/reportes/desempeno/distribucion/", "/api/exportaciones/reportes/distribucion/xlsx/"),
            ("exentos", "/api/reportes/desempeno/exentos/", "/api/exportaciones/reportes/exentos/xlsx/"),
            ("docentes", "/api/reportes/desempeno/docentes/", "/api/exportaciones/reportes/desempeno-docente/xlsx/"),
            ("cohorte", "/api/reportes/desempeno/cohorte/", "/api/exportaciones/reportes/desempeno-cohorte/xlsx/"),
            ("reprobados-nominal", "/api/reportes/desempeno/reprobados-nominal/", "/api/exportaciones/reportes/reprobados-nominal/xlsx/"),
            ("cuadro-aprovechamiento", "/api/reportes/desempeno/cuadro-aprovechamiento/", "/api/exportaciones/reportes/cuadro-aprovechamiento/xlsx/"),
        ]

        self.client.force_login(self.estadistica)
        for slug, json_url, xlsx_url in endpoints:
            json_response = self.client.get(json_url)
            self.assertEqual(json_response.status_code, 200, slug)
            self.assertIn("total", json_response.json(), slug)
            self.assertIn("columnas", json_response.json(), slug)
            self.assertIn("resumen", json_response.json(), slug)

            xlsx_response = self.client.get(xlsx_url)
            self.assert_xlsx_response_desempeno(xlsx_response)
            registro = RegistroExportacion.objects.get(id=xlsx_response["X-Registro-Exportacion-Id"])
            self.assertEqual(registro.estado, RegistroExportacion.ESTADO_GENERADA)
            self.assertEqual(registro.formato, RegistroExportacion.FORMATO_XLSX)

    def test_reportes_desempeno_bloquean_anonimo_discente_y_docente(self):
        anonimo = self.client.get("/api/reportes/desempeno/aprobados-reprobados/")
        self.assertEqual(anonimo.status_code, 401)

        self.client.force_login(self.discentes[0].usuario)
        discente_json = self.client.get("/api/reportes/desempeno/aprobados-reprobados/")
        discente_xlsx = self.client.get("/api/exportaciones/reportes/aprobados-reprobados/xlsx/")
        self.assertEqual(discente_json.status_code, 403)
        self.assertEqual(discente_xlsx.status_code, 403)

        self.client.force_login(self.docente)
        docente_json = self.client.get("/api/reportes/desempeno/aprobados-reprobados/")
        docente_xlsx = self.client.get("/api/exportaciones/reportes/aprobados-reprobados/xlsx/")
        self.assertEqual(docente_json.status_code, 403)
        self.assertEqual(docente_xlsx.status_code, 403)

    def test_reportes_desempeno_jefatura_carrera_filtra_por_ambito(self):
        self.client.force_login(self.jefe_carrera)
        response = self.client.get("/api/reportes/desempeno/aprobados-reprobados/")

        self.assertEqual(response.status_code, 200)
        carreras = {item["carrera"] for item in response.json()["items"]}
        self.assertEqual(carreras, {self.carrera.clave})

    def test_reportes_nominales_no_exponen_matricula(self):
        self.client.force_login(self.estadistica)
        response = self.client.get("/api/reportes/desempeno/reprobados-nominal/")

        self.assertEqual(response.status_code, 200)
        serializado = json.dumps(response.json(), ensure_ascii=False)
        self.assertIn("Discente 3", serializado)
        self.assertNotIn(self.discentes[2].matricula, serializado)
        self.assertNotIn("matricula", serializado.lower())

    def test_reporte_exentos_usa_exencion_de_examen_final(self):
        self.marcar_exento_final()

        self.client.force_login(self.estadistica)
        response = self.client.get("/api/reportes/desempeno/exentos/")

        self.assertEqual(response.status_code, 200)
        payload = json.dumps(response.json(), ensure_ascii=False)
        self.assertIn("Examen final", payload)
        self.assertIn("6.6", payload)

    def test_cuadro_aprovechamiento_clasifica_rangos(self):
        for inscripcion, calificacion in zip(self.inscripciones, [Decimal("9.80"), Decimal("9.30"), Decimal("8.50")]):
            inscripcion.calificacion_final = calificacion
            inscripcion.codigo_resultado_oficial = CatalogoResultadoAcademico.CLAVE_APROBADO
            inscripcion.save(update_fields=["calificacion_final", "codigo_resultado_oficial"])

        self.client.force_login(self.estadistica)
        response = self.client.get("/api/reportes/desempeno/cuadro-aprovechamiento/")

        self.assertEqual(response.status_code, 200)
        payload = json.dumps(response.json(), ensure_ascii=False)
        self.assertIn("Excelente aprovechamiento", payload)
        self.assertIn("Alto aprovechamiento academico", payload)
        self.assertIn("Buen aprovechamiento academico", payload)

    def test_exportacion_reporte_desempeno_fallida_registra_fallida(self):
        self.client.force_login(self.estadistica)
        with patch("reportes.reportes_desempeno.generar_reporte_xlsx", side_effect=RuntimeError("fallo desempeno")):
            response = self.client.get("/api/exportaciones/reportes/aprobados-reprobados/xlsx/")

        self.assertEqual(response.status_code, 500)
        registro = RegistroExportacion.objects.latest("id")
        self.assertEqual(registro.tipo_documento, RegistroExportacion.TIPO_REPORTE_APROBADOS_REPROBADOS)
        self.assertEqual(registro.estado, RegistroExportacion.ESTADO_FALLIDA)
        self.assertIn("fallo desempeno", registro.mensaje_error)

    def test_exportacion_reporte_desempeno_nombre_seguro_y_filtros_sanitizados(self):
        self.client.force_login(self.estadistica)
        response = self.client.get(
            "/api/exportaciones/reportes/aprobados-reprobados/xlsx/",
            {"periodo": self.periodo.clave, "carrera": self.carrera.clave, "password": "secreto"},
        )

        self.assert_xlsx_response_desempeno(response)
        filename = response["Content-Disposition"].split('filename="', 1)[1].rstrip('"')
        registro = RegistroExportacion.objects.get(id=response["X-Registro-Exportacion-Id"])
        self.assertNotIn(" ", filename)
        self.assertNotIn(self.discentes[0].matricula, filename)
        self.assertEqual(registro.filtros_json, {"periodo": self.periodo.clave, "carrera": self.carrera.clave})

    def preparar_trayectoria_reportes(self):
        if not getattr(self, "_trayectoria_preparada", False):
            registrar_extraordinario(
                self.inscripciones[2],
                Decimal("7.0"),
                timezone.localdate(),
                self.estadistica,
            )
            registrar_baja_temporal(
                self.discentes[0],
                fecha_inicio=timezone.localdate(),
                periodo=self.periodo,
                motivo="Baja temporal de prueba",
                registrado_por=self.estadistica,
                inscripcion_materia=self.inscripciones[0],
            )
            registrar_reingreso(
                self.discentes[0],
                fecha_inicio=timezone.localdate(),
                periodo=self.periodo,
                motivo="Reingreso de prueba",
                registrado_por=self.estadistica,
            )
            EventoSituacionAcademica.objects.create(
                discente=self.discentes[1],
                situacion=obtener_situacion(CatalogoSituacionAcademica.CLAVE_BAJA_DEFINITIVA),
                fecha_inicio=timezone.localdate(),
                periodo=self.periodo,
                motivo="Baja definitiva de prueba",
                registrado_por=self.estadistica,
            )
            self.discentes[1].situacion_actual = Discente.SITUACION_BAJA_DEFINITIVA
            self.discentes[1].save(update_fields=["situacion_actual"])
            self.movimiento_discente = self.crear_discente(9)
            self.grupo_destino = GrupoAcademico.objects.create(
                clave_grupo="R9BG9",
                antiguedad=self.antiguedad,
                periodo=self.periodo,
                semestre_numero=1,
            )
            self.movimiento = MovimientoAcademico.objects.create(
                discente=self.movimiento_discente,
                periodo=self.periodo,
                tipo_movimiento=MovimientoAcademico.CAMBIO_GRUPO,
                grupo_origen=self.grupo,
                grupo_destino=self.grupo_destino,
                observaciones="Cambio de grupo de prueba",
            )
            self.discentes[2].situacion_actual = Discente.SITUACION_EGRESADO
            self.discentes[2].save(update_fields=["situacion_actual"])
            EventoSituacionAcademica.objects.create(
                discente=self.discentes[2],
                situacion=obtener_situacion(CatalogoSituacionAcademica.CLAVE_EGRESADO),
                fecha_inicio=timezone.localdate(),
                periodo=self.periodo,
                motivo="Egreso de prueba",
                registrado_por=self.estadistica,
            )
            self._trayectoria_preparada = True

    def assert_xlsx_response_trayectoria(self, response):
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(response.content.startswith(b"PK"))
        self.assertIn("X-Registro-Exportacion-Id", response)

    def test_reportes_trayectoria_json_y_xlsx_como_estadistica(self):
        self.preparar_trayectoria_reportes()
        endpoints = [
            ("extraordinarios", "/api/reportes/situacion/extraordinarios/", "/api/exportaciones/reportes/extraordinarios/xlsx/"),
            ("situacion", "/api/reportes/situacion/actual/", "/api/exportaciones/reportes/situacion-actual/xlsx/"),
            ("bajas-temporales", "/api/reportes/situacion/bajas-temporales/", "/api/exportaciones/reportes/bajas-temporales/xlsx/"),
            ("bajas-definitivas", "/api/reportes/situacion/bajas-definitivas/", "/api/exportaciones/reportes/bajas-definitivas/xlsx/"),
            ("reingresos", "/api/reportes/situacion/reingresos/", "/api/exportaciones/reportes/reingresos/xlsx/"),
            ("egresables", "/api/reportes/situacion/egresables/", "/api/exportaciones/reportes/egresables/xlsx/"),
            ("situacion-agregado", "/api/reportes/situacion/agregado/", "/api/exportaciones/reportes/situacion-agregado/xlsx/"),
            ("movimientos", "/api/reportes/movimientos/", "/api/exportaciones/reportes/movimientos-academicos/xlsx/"),
            ("cambios-grupo", "/api/reportes/movimientos/cambios-grupo/", "/api/exportaciones/reportes/cambios-grupo/xlsx/"),
            ("historial-interno", "/api/reportes/historial-interno/", "/api/exportaciones/reportes/historial-interno/xlsx/"),
        ]
        self.client.force_login(self.estadistica)
        for slug, json_url, xlsx_url in endpoints:
            json_response = self.client.get(json_url)
            self.assertEqual(json_response.status_code, 200, slug)
            self.assertIn("total", json_response.json(), slug)
            self.assertIn("columnas", json_response.json(), slug)
            self.assertIn("resumen", json_response.json(), slug)

            xlsx_response = self.client.get(xlsx_url)
            self.assert_xlsx_response_trayectoria(xlsx_response)
            registro = RegistroExportacion.objects.get(id=xlsx_response["X-Registro-Exportacion-Id"])
            self.assertEqual(registro.estado, RegistroExportacion.ESTADO_GENERADA)
            self.assertEqual(registro.formato, RegistroExportacion.FORMATO_XLSX)

    def test_historial_interno_por_discente_exporta_y_conserva_ordinario_con_ee(self):
        self.preparar_trayectoria_reportes()
        self.client.force_login(self.estadistica)
        response = self.client.get(f"/api/exportaciones/reportes/historial-interno/{self.discentes[2].id}/xlsx/")

        self.assert_xlsx_response_trayectoria(response)
        registro = RegistroExportacion.objects.get(id=response["X-Registro-Exportacion-Id"])
        self.assertEqual(registro.tipo_documento, RegistroExportacion.TIPO_HISTORIAL_ACADEMICO)
        valores = self.valores_xlsx(response)
        texto = " ".join(str(valor) for valor in valores)
        self.assertIn("Conserva evidencia ordinaria previa a EE", texto)
        self.assertIn("5.95", texto)
        self.assertIn("7.0", texto)
        self.assertNotIn(self.discentes[2].matricula, texto)

    def test_reportes_trayectoria_bloquean_anonimo_discente_y_docente(self):
        anonimo = self.client.get("/api/reportes/situacion/actual/")
        self.assertEqual(anonimo.status_code, 401)

        self.client.force_login(self.discentes[0].usuario)
        discente_json = self.client.get("/api/reportes/situacion/actual/")
        discente_xlsx = self.client.get("/api/exportaciones/reportes/situacion-actual/xlsx/")
        self.assertEqual(discente_json.status_code, 403)
        self.assertEqual(discente_xlsx.status_code, 403)

        self.client.force_login(self.docente)
        docente_json = self.client.get("/api/reportes/situacion/actual/")
        docente_xlsx = self.client.get("/api/exportaciones/reportes/situacion-actual/xlsx/")
        self.assertEqual(docente_json.status_code, 403)
        self.assertEqual(docente_xlsx.status_code, 403)

    def test_reportes_trayectoria_jefatura_carrera_filtra_por_ambito(self):
        self.preparar_trayectoria_reportes()
        acta_externa = self.crear_acta_externa_operativa()
        grupo_externo = acta_externa.asignacion_docente.grupo_academico
        plan_externo = acta_externa.asignacion_docente.programa_asignatura.plan_estudios
        usuario_externo = self.crear_usuario("discenteexterno9i", "DISCENTE", nombre="Discente Externo 9I")
        discente_externo = Discente.objects.create(
            usuario=usuario_externo,
            matricula="R9IEXT001",
            plan_estudios=plan_externo,
            antiguedad=grupo_externo.antiguedad,
        )
        AdscripcionGrupo.objects.create(discente=discente_externo, grupo_academico=grupo_externo)
        EventoSituacionAcademica.objects.create(
            discente=discente_externo,
            situacion=obtener_situacion(CatalogoSituacionAcademica.CLAVE_BAJA_TEMPORAL),
            fecha_inicio=timezone.localdate(),
            motivo="Evento externo",
            registrado_por=self.estadistica,
        )

        self.client.force_login(self.jefe_carrera)
        response = self.client.get("/api/reportes/situacion/actual/")

        self.assertEqual(response.status_code, 200)
        ids = {item["discente_id"] for item in response.json()["items"]}
        self.assertIn(self.discentes[0].id, ids)
        self.assertNotIn(discente_externo.id, ids)

    def test_reportes_trayectoria_no_exponen_matricula_y_agregado_no_muestra_nombres(self):
        self.preparar_trayectoria_reportes()
        self.client.force_login(self.estadistica)
        nominal = self.client.get("/api/reportes/situacion/actual/")
        agregado = self.client.get("/api/reportes/situacion/agregado/")

        self.assertEqual(nominal.status_code, 200)
        self.assertEqual(agregado.status_code, 200)
        nominal_texto = json.dumps(nominal.json(), ensure_ascii=False)
        agregado_items = agregado.json()["items"]
        agregado_texto = json.dumps(agregado_items, ensure_ascii=False)
        self.assertIn("Discente 1", nominal_texto)
        self.assertNotIn(self.discentes[0].matricula, nominal_texto)
        self.assertNotIn("matricula", nominal_texto.lower())
        self.assertNotIn("Discente 1", agregado_texto)
        self.assertTrue(all("nombre" not in item and "nombre_discente" not in item for item in agregado_items))

    def test_extraordinario_aprobado_muestra_marca_ee_y_baja_temporal_abierta(self):
        self.preparar_trayectoria_reportes()
        registrar_baja_temporal(
            self.discentes[2],
            fecha_inicio=timezone.localdate(),
            periodo=self.periodo,
            motivo="Baja abierta de prueba",
            registrado_por=self.estadistica,
        )

        self.client.force_login(self.estadistica)
        extraordinarios = self.client.get("/api/reportes/situacion/extraordinarios/")
        bajas = self.client.get("/api/reportes/situacion/bajas-temporales/")

        self.assertEqual(extraordinarios.status_code, 200)
        self.assertIn("EE", json.dumps(extraordinarios.json(), ensure_ascii=False))
        self.assertEqual(bajas.status_code, 200)
        self.assertIn("Si", json.dumps(bajas.json(), ensure_ascii=False))

    def test_reingreso_y_movimientos_se_reportan_sin_modificar_datos(self):
        self.preparar_trayectoria_reportes()
        movimiento_id = self.movimiento.id
        adscripciones_antes = AdscripcionGrupo.objects.count()
        inscripciones_antes = InscripcionMateria.objects.count()

        self.client.force_login(self.estadistica)
        reingresos = self.client.get("/api/reportes/situacion/reingresos/")
        movimientos = self.client.get("/api/reportes/movimientos/cambios-grupo/")

        self.assertEqual(reingresos.status_code, 200)
        self.assertIn("Reingreso de prueba", json.dumps(reingresos.json(), ensure_ascii=False))
        self.assertEqual(movimientos.status_code, 200)
        self.assertIn(movimiento_id, {item["movimiento_id"] for item in movimientos.json()["items"]})
        self.assertEqual(AdscripcionGrupo.objects.count(), adscripciones_antes)
        self.assertEqual(InscripcionMateria.objects.count(), inscripciones_antes)

    def test_exportacion_reporte_trayectoria_fallida_registra_fallida(self):
        self.client.force_login(self.estadistica)
        with patch("reportes.reportes_trayectoria.generar_reporte_xlsx", side_effect=RuntimeError("fallo trayectoria")):
            response = self.client.get("/api/exportaciones/reportes/situacion-actual/xlsx/")

        self.assertEqual(response.status_code, 500)
        registro = RegistroExportacion.objects.latest("id")
        self.assertEqual(registro.tipo_documento, RegistroExportacion.TIPO_REPORTE_SITUACION_ACADEMICA)
        self.assertEqual(registro.estado, RegistroExportacion.ESTADO_FALLIDA)
        self.assertIn("fallo trayectoria", registro.mensaje_error)

    def test_exportacion_reporte_trayectoria_nombre_seguro_y_filtros_sanitizados(self):
        self.client.force_login(self.estadistica)
        response = self.client.get(
            "/api/exportaciones/reportes/situacion-actual/xlsx/",
            {"periodo": self.periodo.clave, "carrera": self.carrera.clave, "password": "secreto"},
        )

        self.assert_xlsx_response_trayectoria(response)
        filename = response["Content-Disposition"].split('filename="', 1)[1].rstrip('"')
        registro = RegistroExportacion.objects.get(id=response["X-Registro-Exportacion-Id"])
        self.assertNotIn(" ", filename)
        self.assertNotIn(self.discentes[0].matricula, filename)
        self.assertEqual(registro.filtros_json, {"periodo": self.periodo.clave, "carrera": self.carrera.clave})
