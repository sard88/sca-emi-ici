from copy import copy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from .libreoffice_utils import convertir_xlsx_a_pdf as convertir_xlsx_a_pdf_generico

TEMPLATE_KARDEX = Path(__file__).resolve().parents[1] / "templates_xlsx" / "kardex" / "kardex_oficial_template.xlsx"
START_ROW_MATERIAS = 16
MAX_FILAS_TEMPLATE = 40

LEYENDAS_KARDEX = (
    "Escala de calificaciones de 0 a 10 puntos. La mínima aprobatoria es 6.0.\n"
    "EE: Examen Extraordinario. Las materias no numéricas pueden mostrarse como ACREDITADA, APROBADO o EXCEPTUADO.\n"
    "NOTA: Este documento no es válido si presenta raspaduras o enmendaduras."
)


class ExportadorKardexPDF:
    """Genera kárdex PDF convirtiendo el XLSX institucional final con LibreOffice."""

    def generar(self, contexto):
        xlsx_bytes = ExportadorKardexExcelInterno().generar(contexto)
        return convertir_xlsx_a_pdf(xlsx_bytes)


class ExportadorKardexExcelInterno:
    """Genera XLSX temporal de kárdex. No se expone como descarga pública en 9C."""

    def generar(self, contexto):
        if not TEMPLATE_KARDEX.exists():
            raise FileNotFoundError(f"No se encontró la plantilla XLSX de kárdex: {TEMPLATE_KARDEX}")
        wb = load_workbook(TEMPLATE_KARDEX)
        ws = wb.active
        self._poblar_encabezado(ws, contexto)
        self._poblar_materias(ws, contexto)
        self._poblar_promedios_y_certificacion(ws, contexto)
        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    def _poblar_encabezado(self, ws, contexto):
        self._set(ws, "C7", contexto.nombre_discente)
        self._set(ws, "G7", contexto.situacion_actual)
        self._set(ws, "C8", contexto.grado_empleo or "N/A")
        self._set(ws, "G8", contexto.fecha_emision)
        self._set(ws, "C9", f"{contexto.carrera_clave} - {contexto.carrera_oficial or contexto.carrera_nombre}")
        self._set(ws, "C10", contexto.plan_estudios)
        self._set(ws, "G10", contexto.antiguedad)
        self._set(ws, "A12", contexto.lugar_fecha)

    def _poblar_materias(self, ws, contexto):
        materias = [materia for anio in contexto.anios for materia in anio.materias]
        self._expandir_tabla(ws, len(materias))
        self._limpiar_tabla(ws, START_ROW_MATERIAS, max(MAX_FILAS_TEMPLATE, len(materias)))
        for offset, materia in enumerate(materias):
            row = START_ROW_MATERIAS + offset
            valores = [
                materia.anio_label,
                materia.semestre_numero,
                materia.clave_materia,
                materia.nombre_materia,
                materia.calificacion_display,
                materia.calificacion_letra_display,
                materia.marca,
                materia.resultado_display,
            ]
            for col, value in enumerate(valores, start=1):
                self._set(ws, row=row, column=col, value=value)

    def _expandir_tabla(self, ws, total_materias):
        if total_materias <= MAX_FILAS_TEMPLATE:
            return
        extra_rows = total_materias - MAX_FILAS_TEMPLATE
        insert_at = START_ROW_MATERIAS + MAX_FILAS_TEMPLATE
        ws.insert_rows(insert_at, extra_rows)
        template_row = insert_at - 1
        for row in range(insert_at, insert_at + extra_rows):
            ws.row_dimensions[row].height = ws.row_dimensions[template_row].height
            for col in range(1, 9):
                source = ws.cell(template_row, col)
                target = ws.cell(row, col)
                target._style = copy(source._style)
                target.font = copy(source.font)
                target.fill = copy(source.fill)
                target.border = copy(source.border)
                target.alignment = copy(source.alignment)
        ws.print_area = f"A1:H{75 + extra_rows}"

    def _limpiar_tabla(self, ws, start_row, rows):
        for row in range(start_row, start_row + rows):
            for col in range(1, 9):
                self._set(ws, row=row, column=col, value="")

    def _poblar_promedios_y_certificacion(self, ws, contexto):
        summary_start = START_ROW_MATERIAS + max(MAX_FILAS_TEMPLATE, sum(len(anio.materias) for anio in contexto.anios)) + 3
        self._set(ws, f"A{summary_start}", "Promedios anuales")
        for offset, anio in enumerate(contexto.anios[:6], start=1):
            row = summary_start + offset
            self._set(ws, f"A{row}", anio.anio_label)
            self._set(ws, f"B{row}", anio.promedio_anual_display)
        self._set(ws, f"E{summary_start}", "Promedio general")
        self._set(ws, f"E{summary_start + 1}", contexto.promedio_general_display)

        leyendas_row = summary_start + 8
        self._set(ws, f"A{leyendas_row}", "Leyendas")
        self._set(ws, f"A{leyendas_row + 1}", LEYENDAS_KARDEX)

        cert_row = leyendas_row + 5
        self._set(ws, f"A{cert_row}", "Certificación")
        self._set(ws, f"A{cert_row + 1}", contexto.certificacion.texto)
        self._set(
            ws,
            f"A{cert_row + 3}",
            f"{contexto.certificacion.autoridad_cargo}\n{contexto.certificacion.autoridad_nombre}\n{contexto.certificacion.unidad_responsable}",
        )
        self._set(ws, f"E{cert_row + 3}", "________________________________\nFirma")
        ws.print_area = f"A1:H{cert_row + 4}"

    def _set(self, ws, cell=None, value=None, *, row=None, column=None):
        target = ws[cell] if cell else ws.cell(row=row, column=column)
        if isinstance(target, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if target.coordinate in merged_range:
                    target = ws.cell(merged_range.min_row, merged_range.min_col)
                    break
        target.value = value


def convertir_xlsx_a_pdf(xlsx_bytes):
    return convertir_xlsx_a_pdf_generico(xlsx_bytes, prefix="sca_kardex_pdf_", stem="kardex")
