from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from reportes.constantes import LEYENDA_ESCALA, LEYENDA_INCONFORMIDAD

TEMPLATES_DIR = Path(__file__).resolve().parents[1] / "templates_xlsx" / "actas"
TEMPLATE_PARCIAL = TEMPLATES_DIR / "acta_evaluacion_parcial_template.xlsx"
TEMPLATE_EVALUACION_FINAL = TEMPLATES_DIR / "acta_evaluacion_final_template.xlsx"
TEMPLATE_CALIFICACION_FINAL = TEMPLATES_DIR / "acta_calificacion_final_template.xlsx"

MAX_FILAS_PARCIAL = 20
MAX_FILAS_FINAL = 17

PARCIAL_COMPONENT_CELLS = ["U15", "Z15", "AE15", "AJ15"]
PARCIAL_ROW_GROUPS = [
    ("A", "B"),
    ("C", "H"),
    ("I", "T"),
    ("U", "Y"),
    ("Z", "AD"),
    ("AE", "AI"),
    ("AJ", "AN"),
    ("AO", "AS"),
    ("AT", "AX"),
]
EVALUACION_FINAL_COMPONENT_CELLS = ["D13", "E13", "F13"]


class ExportadorActaExcel:
    """Genera XLSX desde plantillas institucionales anonimizadas."""

    def generar(self, contexto):
        if contexto.es_calificacion_final:
            wb = self._cargar_template(TEMPLATE_CALIFICACION_FINAL)
            self._poblar_calificacion_final(wb, contexto)
        elif contexto.tipo_documento == "ACTA_EVALUACION_FINAL":
            wb = self._cargar_template(TEMPLATE_EVALUACION_FINAL)
            self._poblar_evaluacion_final(wb, contexto)
        else:
            wb = self._cargar_template(TEMPLATE_PARCIAL)
            self._poblar_parcial(wb, contexto)

        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    def _cargar_template(self, path):
        if not path.exists():
            raise FileNotFoundError(f"No se encontró la plantilla XLSX institucional: {path}")
        return load_workbook(path)

    def _poblar_parcial(self, wb, contexto):
        filas_paginas = self._partir(contexto.filas_corte, MAX_FILAS_PARCIAL)
        sheets = self._preparar_paginas(wb, filas_paginas, "Acta parcial")
        for page_index, (ws, filas) in enumerate(zip(sheets, filas_paginas), start=1):
            ws.title = self._titulo_hoja("Acta parcial", page_index, len(sheets))
            self._encabezado_parcial(ws, contexto)
            self._marca_no_oficial(ws, "A14", contexto)
            self._headers_parcial(ws, contexto)
            self._filas_parcial(ws, filas, page_index)
            self._ocultar_filas_sobrantes(ws, start_row=16, capacidad=MAX_FILAS_PARCIAL, usadas=len(filas))
            self._resumen_parcial(ws, contexto)
            self._firmas_parcial(ws, contexto)
            self._leyenda(ws, "A56")

    def _poblar_evaluacion_final(self, wb, contexto):
        filas_paginas = self._partir(contexto.filas_corte, MAX_FILAS_FINAL)
        sheets = self._preparar_paginas(wb, filas_paginas, "Evaluación final")
        for page_index, (ws, filas) in enumerate(zip(sheets, filas_paginas), start=1):
            ws.title = self._titulo_hoja("Evaluación final", page_index, len(sheets))
            self._encabezado_compacto(ws, contexto, metadata_value_col="F", titulo_evaluacion=contexto.evaluacion)
            self._marca_no_oficial(ws, "A12", contexto)
            self._headers_evaluacion_final(ws, contexto)
            self._filas_evaluacion_final(ws, filas, page_index)
            self._ocultar_filas_sobrantes(ws, start_row=14, capacidad=MAX_FILAS_FINAL, usadas=len(filas))
            self._resumen_compacto(ws, contexto, value_col="H")
            self._firmas_compactas(ws, contexto, derecha_col="D")
            self._leyenda(ws, "A49")

    def _poblar_calificacion_final(self, wb, contexto):
        filas_paginas = self._partir(contexto.filas_final, MAX_FILAS_FINAL)
        sheets = self._preparar_paginas(wb, filas_paginas, "Calificación final")
        for page_index, (ws, filas) in enumerate(zip(sheets, filas_paginas), start=1):
            ws.title = self._titulo_hoja("Calificación final", page_index, len(sheets))
            self._encabezado_compacto(ws, contexto, metadata_value_col="G", titulo_evaluacion="Calificación Final")
            self._marca_no_oficial(ws, "A12", contexto)
            self._filas_calificacion_final(ws, filas, page_index)
            self._ocultar_filas_sobrantes(ws, start_row=14, capacidad=MAX_FILAS_FINAL, usadas=len(filas))
            self._resumen_compacto(ws, contexto, value_col="I")
            self._firmas_compactas(ws, contexto, derecha_col="E")
            self._leyenda(ws, "A49")

    def _preparar_paginas(self, wb, filas_paginas, base_title):
        ws_base = wb.active
        sheets = [ws_base]
        for _ in range(1, len(filas_paginas)):
            sheets.append(wb.copy_worksheet(ws_base))
        for idx, ws in enumerate(sheets, start=1):
            ws.title = self._titulo_hoja(base_title, idx, len(sheets))
        return sheets

    def _partir(self, filas, capacidad):
        filas = list(filas)
        if not filas:
            return [[]]
        return [filas[index:index + capacidad] for index in range(0, len(filas), capacidad)]

    def _titulo_hoja(self, titulo, index, total):
        if total == 1:
            return titulo[:31]
        return f"{titulo} {index}"[:31]

    def _encabezado_parcial(self, ws, contexto):
        self._set(ws, "A5", self._texto_titulo(contexto))
        self._set(ws, "AJ8", contexto.unidad_aprendizaje)
        self._set(ws, "AJ9", contexto.docente)
        self._set(ws, "AJ10", contexto.grupo)
        self._set(ws, "AJ11", contexto.ciclo_escolar)
        self._set(ws, "AJ12", contexto.semestre)
        self._set(ws, "AJ13", contexto.evaluacion)

    def _encabezado_compacto(self, ws, contexto, *, metadata_value_col, titulo_evaluacion):
        self._set(ws, "A5", self._texto_titulo(contexto))
        self._set(ws, f"{metadata_value_col}6", contexto.unidad_aprendizaje)
        self._set(ws, f"{metadata_value_col}7", contexto.docente)
        self._set(ws, f"{metadata_value_col}8", contexto.grupo)
        self._set(ws, f"{metadata_value_col}9", contexto.ciclo_escolar)
        self._set(ws, f"{metadata_value_col}10", contexto.semestre)
        self._set(ws, f"{metadata_value_col}11", titulo_evaluacion)

    def _headers_parcial(self, ws, contexto):
        for cell in PARCIAL_COMPONENT_CELLS:
            self._set(ws, cell, "")
        for cell, componente in zip(PARCIAL_COMPONENT_CELLS, contexto.componentes[:4]):
            self._set(ws, cell, self._encabezado_componente(componente))
        self._set(ws, "AO15", "Calificación\nfinal")
        self._set(ws, "AT15", "Firma de\nconformidad")

    def _headers_evaluacion_final(self, ws, contexto):
        for cell in EVALUACION_FINAL_COMPONENT_CELLS:
            self._set(ws, cell, "")
        for cell, componente in zip(EVALUACION_FINAL_COMPONENT_CELLS, contexto.componentes[:3]):
            self._set(ws, cell, self._encabezado_componente(componente))
        self._set(ws, "G13", "Calificación final")
        self._set(ws, "H13", "Firma de conformidad")

    def _filas_parcial(self, ws, filas, page_index):
        start_row = 16
        self._limpiar_rango(ws, start_row, start_row + MAX_FILAS_PARCIAL - 1, 1, 50)
        offset_global = (page_index - 1) * MAX_FILAS_PARCIAL
        for offset, fila in enumerate(filas):
            row = start_row + offset
            valores = [
                offset_global + offset + 1,
                fila.grado_empleo,
                fila.nombre,
                *self._componentes_normalizados(fila.componentes, 4),
                fila.calificacion_final,
                fila.firma_conformidad,
            ]
            for (start_col, _end_col), value in zip(PARCIAL_ROW_GROUPS, valores):
                self._set(ws, f"{start_col}{row}", self._valor(value))

    def _filas_evaluacion_final(self, ws, filas, page_index):
        start_row = 14
        self._limpiar_rango(ws, start_row, start_row + MAX_FILAS_FINAL - 1, 1, 8)
        offset_global = (page_index - 1) * MAX_FILAS_FINAL
        for offset, fila in enumerate(filas):
            row = start_row + offset
            componentes = self._componentes_normalizados(fila.componentes, 3)
            valores = [
                offset_global + offset + 1,
                fila.grado_empleo,
                fila.nombre,
                *componentes,
                fila.calificacion_final,
                fila.firma_conformidad,
            ]
            for col, value in enumerate(valores, start=1):
                self._set(ws, row=row, column=col, value=self._valor(value))

    def _filas_calificacion_final(self, ws, filas, page_index):
        start_row = 14
        self._limpiar_rango(ws, start_row, start_row + MAX_FILAS_FINAL - 1, 1, 9)
        offset_global = (page_index - 1) * MAX_FILAS_FINAL
        for offset, fila in enumerate(filas):
            row = start_row + offset
            valores = [
                offset_global + offset + 1,
                fila.grado_empleo,
                fila.nombre,
                fila.parcial_1,
                fila.parcial_2,
                fila.parcial_3,
                fila.evaluacion_final,
                fila.calificacion_final,
                fila.firma_conformidad,
            ]
            for col, value in enumerate(valores, start=1):
                self._set(ws, row=row, column=col, value=self._valor(value))

    def _resumen_parcial(self, ws, contexto):
        stats = contexto.estadisticos
        self._set(ws, "A37", f"Probables causas de reprobación: {contexto.causas_reprobacion}")
        self._set(ws, "A39", f"Sugerencias : {contexto.sugerencias}")
        self._set(ws, "AT37", stats.alumnos_reprobados if stats.alumnos_reprobados is not None else "N/A")
        self._set(ws, "AT38", self._valor(stats.media))
        self._set(ws, "AT39", self._valor(stats.moda))
        self._set(ws, "AT40", self._valor(stats.desviacion_estandar))
        if contexto.nota_exencion:
            self._set(ws, "A41", f"Nota: {contexto.nota_exencion}")
        self._set(ws, "A42", contexto.lugar_fecha)

    def _resumen_compacto(self, ws, contexto, *, value_col):
        stats = contexto.estadisticos
        self._set(ws, "A33", f"Probables causas de reprobación: {contexto.causas_reprobacion}")
        self._set(ws, "A36", f"Sugerencias : {contexto.sugerencias}")
        self._set(ws, f"{value_col}32", stats.alumnos_reprobados if stats.alumnos_reprobados is not None else "N/A")
        self._set(ws, f"{value_col}33", self._valor(stats.media))
        self._set(ws, f"{value_col}34", self._valor(stats.moda))
        self._set(ws, f"{value_col}35", self._valor(stats.desviacion_estandar))
        if contexto.nota_exencion:
            self._set(ws, "A38", f"Nota: {contexto.nota_exencion}")
        date_cell = "H39" if value_col == "H" else "I39"
        self._set(ws, date_cell, contexto.lugar_fecha)

    def _firmas_parcial(self, ws, contexto):
        evaluo, reviso, vo_bo = self._firmas(contexto)
        self._set(ws, "A44", f"{evaluo.etiqueta}:")
        self._set(ws, "A45", evaluo.cargo)
        self._set(ws, "A47", evaluo.nombre)
        self._set(ws, "A48", evaluo.cedula)
        self._set(ws, "Z44", f"{reviso.etiqueta}:")
        self._set(ws, "Z45", reviso.cargo)
        self._set(ws, "Z47", reviso.nombre)
        self._set(ws, "A50", vo_bo.etiqueta)
        self._set(ws, "A51", vo_bo.cargo)
        self._set(ws, "A53", vo_bo.nombre)

    def _firmas_compactas(self, ws, contexto, *, derecha_col):
        evaluo, reviso, _vo_bo = self._firmas(contexto)
        self._set(ws, "B42", f"{evaluo.etiqueta}:")
        self._set(ws, "B43", evaluo.cargo)
        self._set(ws, "B45", evaluo.nombre)
        self._set(ws, "B46", evaluo.cedula)
        self._set(ws, f"{derecha_col}42", f"{reviso.etiqueta}:")
        self._set(ws, f"{derecha_col}43", reviso.cargo)
        self._set(ws, f"{derecha_col}45", reviso.nombre)

    def _leyenda(self, ws, cell):
        self._set(ws, cell, f"{LEYENDA_ESCALA}\n{LEYENDA_INCONFORMIDAD}")

    def _marca_no_oficial(self, ws, cell, contexto):
        self._set(ws, cell, contexto.marca_no_oficial or "")

    def _firmas(self, contexto):
        firmas = list(contexto.firmas)
        while len(firmas) < 3:
            firmas.append(type(firmas[0])(etiqueta="Pendiente") if firmas else None)
        return firmas[:3]

    def _componentes_normalizados(self, componentes, total):
        valores = list(componentes[:total])
        while len(valores) < total:
            valores.append("")
        return valores

    def _encabezado_componente(self, componente):
        return componente.encabezado.replace(" ", "\n", 1) if " " in componente.encabezado else componente.encabezado

    def _texto_titulo(self, contexto):
        return (
            "Acta de calificaciones del personal de Discentes de la carrera de "
            f"{contexto.carrera.upper()}, con anotación de Número de lista, Empleo, Nombre, "
            "Calificación y Firma de conformidad."
        )

    def _valor(self, value):
        if value == "EXENTO":
            return value
        if hasattr(value, "quantize"):
            return float(value)
        return value

    def _set(self, ws, coordinate=None, value=None, *, row=None, column=None):
        cell = ws[coordinate] if coordinate else ws.cell(row=row, column=column)
        if isinstance(cell, MergedCell):
            cell = self._celda_maestra(ws, cell.coordinate)
        cell.value = value
        return cell

    def _celda_maestra(self, ws, coordinate):
        for merged_range in ws.merged_cells.ranges:
            if coordinate in merged_range:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        return ws[coordinate]

    def _limpiar_rango(self, ws, min_row, max_row, min_col, max_col):
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.value = None

    def _ocultar_filas_sobrantes(self, ws, *, start_row, capacidad, usadas):
        for row in range(start_row, start_row + capacidad):
            row_dimension = ws.row_dimensions[row]
            row_dimension.hidden = row >= start_row + usadas
