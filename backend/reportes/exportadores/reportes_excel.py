from dataclasses import dataclass
from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass(frozen=True)
class ReporteSheet:
    titulo: str
    columnas: list[dict]
    filas: list[dict]


def generar_reporte_xlsx(*, titulo: str, filtros: dict, sheets: list[ReporteSheet]) -> bytes:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet in sheets:
        _agregar_hoja(wb, titulo=titulo, filtros=filtros, sheet=sheet)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _agregar_hoja(wb, *, titulo: str, filtros: dict, sheet: ReporteSheet):
    ws = wb.create_sheet(_safe_sheet_title(sheet.titulo))
    max_col = max(1, len(sheet.columnas))
    last_col = get_column_letter(max_col)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = "Sistema de Control Académico EMI - ICI"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="611232")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws["A2"] = titulo
    ws["A2"].font = Font(bold=True, size=12, color="10372E")
    ws["A2"].alignment = Alignment(horizontal="center")

    filtros_texto = ", ".join(f"{key}: {value}" for key, value in filtros.items() if value not in ("", None, [])) or "Sin filtros"
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
    ws["A3"] = f"Filtros aplicados: {filtros_texto}"
    ws["A3"].font = Font(italic=True, color="5F6764")

    header_row = 5
    thin = Side(style="thin", color="D8C5A7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="235B4E")

    for col_idx, col in enumerate(sheet.columnas, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col["label"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, fila in enumerate(sheet.filas, start=header_row + 1):
        for col_idx, col in enumerate(sheet.columnas, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=fila.get(col["key"], ""))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{last_col}{max(header_row, header_row + len(sheet.filas))}"

    for col_idx, col in enumerate(sheet.columnas, start=1):
        values = [str(row.get(col["key"], "")) for row in sheet.filas[:100]]
        width = max([len(str(col["label"]))] + [len(value) for value in values] + [10])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 42)

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4


def _safe_sheet_title(value: str) -> str:
    value = re.sub(r"[\[\]\*:/\\?]", " ", value).strip() or "Reporte"
    return value[:31]
